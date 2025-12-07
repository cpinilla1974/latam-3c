"""
Generador de Informes Excel
Piloto IA - FICEM BD

Genera informes de benchmarking en formato Excel con:
- Múltiples hojas con datos y análisis
- Tablas dinámicas
- Gráficos integrados
"""

import os
from datetime import datetime
from typing import Dict, List, Optional
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, ScatterChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

# RAG y SQL
from ai_modules.rag.rag_chain import BenchmarkingChain
from ai_modules.rag.sql_tool import SQLTool


class BenchmarkingReportExcel:
    """
    Generador de informes de benchmarking en Excel.
    """

    def __init__(
        self,
        llm_model: str = "qwen2.5:3b",
        use_claude: bool = False,
        output_dir: str = "./reports"
    ):
        """
        Inicializa el generador de informes.

        Args:
            llm_model: Modelo LLM a usar
            use_claude: Si True, usa Claude API
            output_dir: Directorio de salida para Excel
        """
        self.llm_model = llm_model
        self.use_claude = use_claude
        self.output_dir = output_dir

        # Crear directorio de salida si no existe
        os.makedirs(output_dir, exist_ok=True)

        # Inicializar RAG chain
        self.rag = BenchmarkingChain(
            llm_model=llm_model,
            temperature=0.1,
            top_k=5,
            use_claude=use_claude
        )

        # Inicializar SQL tool
        self.sql_tool = SQLTool()

    def _apply_header_style(self, ws, row_num: int, col_start: int, col_end: int):
        """
        Aplica estilo a los encabezados.

        Args:
            ws: Worksheet de openpyxl
            row_num: Número de fila del encabezado
            col_start: Columna inicial
            col_end: Columna final
        """
        header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col in range(col_start, col_end + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

    def _apply_data_style(self, ws, row_start: int, row_end: int, col_start: int, col_end: int):
        """
        Aplica estilo a los datos.

        Args:
            ws: Worksheet
            row_start: Fila inicial
            row_end: Fila final
            col_start: Columna inicial
            col_end: Columna final
        """
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in range(row_start, row_end + 1):
            for col in range(col_start, col_end + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")

                # Alternar colores
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

    def generate_company_report(
        self,
        compania: str,
        año: Optional[int] = None,
        benchmark_type: str = "gcca"
    ) -> str:
        """
        Genera informe de benchmarking en Excel.

        Args:
            compania: Nombre de la compañía
            año: Año (opcional)
            benchmark_type: Tipo de benchmark

        Returns:
            Ruta al archivo Excel generado
        """
        print(f"\n📊 Generando informe Excel para {compania}")

        # 1. Obtener datos de la empresa
        result = self.sql_tool.get_huella_promedio_compania(compania, año)
        
        if not result["success"] or not result["rows"]:
            raise ValueError(f"No se encontraron datos para {compania}")

        company_data = result["rows"][0]

        # 2. Obtener análisis de IA
        print("🤖 Generando análisis con IA...")
        ai_analysis = self.rag.compare_with_benchmark(
            company_data=company_data,
            benchmark_type=benchmark_type
        )

        # 3. Obtener distribución de productos
        query_productos = f"""
        SELECT
            resistencia,
            COUNT(*) as num_remitos,
            AVG(huella_co2) as huella_promedio,
            AVG(contenido_cemento) as cemento_promedio,
            SUM(volumen) as volumen_total
        FROM remitos_concretos
        WHERE LOWER(compania) = LOWER('{compania}')
        {'AND año = ' + str(año) if año else ''}
        GROUP BY resistencia
        ORDER BY volumen_total DESC
        """
        result_productos = self.sql_tool.execute_query(query_productos)
        productos_data = result_productos["rows"] if result_productos["success"] else []

        # 4. Obtener evolución temporal (si no hay año específico)
        if not año:
            query_temporal = f"""
            SELECT
                año,
                COUNT(*) as num_remitos,
                AVG(huella_co2) as huella_promedio,
                AVG(resistencia) as resistencia_promedio,
                SUM(volumen) as volumen_total
            FROM remitos_concretos
            WHERE LOWER(compania) = LOWER('{compania}')
            GROUP BY año
            ORDER BY año
            """
            result_temporal = self.sql_tool.execute_query(query_temporal)
            temporal_data = result_temporal["rows"] if result_temporal["success"] else []
        else:
            temporal_data = []

        # 5. Crear archivo Excel
        year_str = str(año) if año else "historico"
        filename = f"benchmarking_{compania}_{year_str}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.output_dir, filename)

        wb = Workbook()

        # Hoja 1: Resumen Ejecutivo
        ws_summary = wb.active
        ws_summary.title = "Resumen"

        # Título
        ws_summary['A1'] = f"Informe de Benchmarking - {compania}"
        ws_summary['A1'].font = Font(bold=True, size=16, color="2E86AB")
        ws_summary['A2'] = f"Período: {año if año else 'Histórico'}"
        ws_summary['A2'].font = Font(italic=True, size=12)
        ws_summary['A3'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"

        # Métricas principales
        row = 5
        ws_summary[f'A{row}'] = "Métrica"
        ws_summary[f'B{row}'] = "Valor"
        self._apply_header_style(ws_summary, row, 1, 2)

        row += 1
        metrics = [
            ("Número de remitos", f"{company_data.get('num_remitos', 0):,}"),
            ("Huella promedio", f"{company_data.get('huella_promedio', 0):.2f} kg CO₂/m³"),
            ("Resistencia promedio", f"{company_data.get('resistencia_promedio', 0):.1f} MPa"),
            ("Contenido cemento", f"{company_data.get('cemento_promedio', 0):.0f} kg/m³"),
            ("Volumen total", f"{company_data.get('volumen_total', 0):,.0f} m³")
        ]

        for metric, value in metrics:
            ws_summary[f'A{row}'] = metric
            ws_summary[f'B{row}'] = value
            row += 1

        self._apply_data_style(ws_summary, 6, row-1, 1, 2)

        # Ajustar anchos de columnas
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 20

        # Hoja 2: Análisis de IA
        ws_ai = wb.create_sheet("Análisis IA")
        ws_ai['A1'] = "Análisis Comparativo Generado por IA"
        ws_ai['A1'].font = Font(bold=True, size=14, color="2E86AB")
        ws_ai.column_dimensions['A'].width = 100

        # Escribir análisis (línea por línea)
        ai_text = ai_analysis.get('answer', 'No disponible')
        row = 3
        for line in ai_text.split('\n'):
            if line.strip():
                ws_ai[f'A{row}'] = line.strip()
                ws_ai[f'A{row}'].alignment = Alignment(wrap_text=True, vertical="top")
                row += 1

        # Hoja 3: Productos por Resistencia
        ws_products = wb.create_sheet("Productos")
        
        if productos_data:
            df_productos = pd.DataFrame(productos_data)
            df_productos.columns = ['Resistencia (MPa)', 'Remitos', 'Huella CO₂', 'Cemento (kg/m³)', 'Volumen (m³)']
            
            # Escribir dataframe
            for r_idx, row in enumerate(dataframe_to_rows(df_productos, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws_products.cell(row=r_idx, column=c_idx, value=value)

            # Aplicar estilos
            self._apply_header_style(ws_products, 1, 1, len(df_productos.columns))
            self._apply_data_style(ws_products, 2, len(df_productos)+1, 1, len(df_productos.columns))

            # Ajustar anchos
            for col_idx, col in enumerate(['B', 'C', 'D', 'E', 'F'], 1):
                ws_products.column_dimensions[col].width = 15

            # Agregar gráfico de barras
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Huella CO₂ por Resistencia"
            chart.y_axis.title = "Huella CO₂ (kg/m³)"
            chart.x_axis.title = "Resistencia (MPa)"

            data = Reference(ws_products, min_col=3, min_row=1, max_row=min(len(productos_data)+1, 15))
            cats = Reference(ws_products, min_col=1, min_row=2, max_row=min(len(productos_data)+1, 15))
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 10
            chart.width = 20

            ws_products.add_chart(chart, "H2")

        # Hoja 4: Evolución Temporal (solo si hay datos históricos)
        if temporal_data:
            ws_temporal = wb.create_sheet("Evolución Temporal")
            
            df_temporal = pd.DataFrame(temporal_data)
            df_temporal.columns = ['Año', 'Remitos', 'Huella CO₂', 'Resistencia', 'Volumen']
            
            # Escribir dataframe
            for r_idx, row in enumerate(dataframe_to_rows(df_temporal, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws_temporal.cell(row=r_idx, column=c_idx, value=value)

            # Aplicar estilos
            self._apply_header_style(ws_temporal, 1, 1, len(df_temporal.columns))
            self._apply_data_style(ws_temporal, 2, len(df_temporal)+1, 1, len(df_temporal.columns))

            # Gráfico de línea
            from openpyxl.chart import LineChart
            chart = LineChart()
            chart.title = "Evolución de Huella CO₂"
            chart.y_axis.title = "Huella CO₂ (kg/m³)"
            chart.x_axis.title = "Año"

            data = Reference(ws_temporal, min_col=3, min_row=1, max_row=len(temporal_data)+1)
            cats = Reference(ws_temporal, min_col=1, min_row=2, max_row=len(temporal_data)+1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 10
            chart.width = 20

            ws_temporal.add_chart(chart, "G2")

        # Guardar archivo
        wb.save(filepath)

        print(f"✅ Informe Excel generado: {filepath}")
        return filepath


# Ejemplo de uso
if __name__ == "__main__":
    generator = BenchmarkingReportExcel(
        llm_model="qwen2.5:3b",
        use_claude=False,
        output_dir="./reports"
    )

    # Generar informe para MZMA 2024
    excel_path = generator.generate_company_report(
        compania="MZMA",
        año=2024,
        benchmark_type="gcca"
    )

    print(f"\n📄 Excel generado: {excel_path}")
