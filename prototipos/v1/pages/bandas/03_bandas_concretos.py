import streamlit as st
import pandas as pd
import os
import json
import matplotlib.pyplot as plt
import numpy as np
from matplotlib.colors import LinearSegmentedColormap, Normalize
from matplotlib.cm import ScalarMappable
from database.connection import get_connection
import plotly.graph_objects as go
import plotly.express as px
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from io import BytesIO

def cargar_bandas(json_path):
    with open(json_path, 'r', encoding='utf-8') as f:
        bandas = json.load(f)
    # Asegurar que las claves de resistencia sean enteros
    bandas = {k: {int(r): v for r, v in d.items()} for k, d in bandas.items()}
    return bandas

def generar_excel_analisis(df_csv, bandas):
    """
    Genera un archivo Excel con 4 hojas:
    - Índice con comparativa
    - Todas las resistencias
    - Solo exactos (20,25,30,35,40,45,50)
    - Rangos ±2 MPa
    """
    resistencias_estandar = [20, 25, 30, 35, 40, 45, 50]

    # Crear workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remover hoja por defecto

    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # HOJA 1: Índice con comparativa
    ws_indice = wb.create_sheet("Índice")

    # Título
    ws_indice['A1'] = "Análisis Comparativo de Opciones de Filtrado"
    ws_indice['A1'].font = Font(size=14, bold=True)

    # Definiciones
    ws_indice['A3'] = "Definiciones:"
    ws_indice['A3'].font = Font(bold=True)

    ws_indice['A4'] = "1. Todas las Resistencias"
    ws_indice['B4'] = "Incluye todos los valores de resistencia sin filtrado"

    ws_indice['A5'] = "2. Solo Exactos"
    ws_indice['B5'] = f"Solo resistencias exactas: {', '.join(map(str, resistencias_estandar))} MPa"

    ws_indice['A6'] = "3. Rangos ±2 MPa"
    ws_indice['B6'] = "Agrupa resistencias en rangos de ±2 MPa alrededor de valores estándar"

    # Calcular estadísticas para cada opción
    ws_indice['A8'] = "Comparativa de Métricas"
    ws_indice['A8'].font = Font(bold=True, size=12)

    # Encabezados
    headers = ['Métrica', 'Todas', 'Solo Exactos', 'Rangos ±2']
    for col, header in enumerate(headers, start=1):
        cell = ws_indice.cell(row=9, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Calcular métricas
    # Opción 1: Todas
    df_todas = df_csv.copy()
    huella_min_todas = df_todas['huella_co2'].min()
    huella_max_todas = df_todas['huella_co2'].max()
    huella_prom_todas = df_todas['huella_co2'].mean()
    volumen_todas = df_todas['volumen'].sum() if 'volumen' in df_todas.columns else 0

    # Opción 2: Solo exactos
    df_exactos = df_csv[df_csv['REST'].isin(resistencias_estandar)]
    huella_min_exactos = df_exactos['huella_co2'].min() if not df_exactos.empty else 0
    huella_max_exactos = df_exactos['huella_co2'].max() if not df_exactos.empty else 0
    huella_prom_exactos = df_exactos['huella_co2'].mean() if not df_exactos.empty else 0
    volumen_exactos = df_exactos['volumen'].sum() if 'volumen' in df_exactos.columns and not df_exactos.empty else 0

    # Opción 3: Rangos
    df_rangos_list = []
    for res_std in resistencias_estandar:
        df_rango = df_csv[(df_csv['REST'] >= res_std - 2) & (df_csv['REST'] <= res_std + 2)]
        if not df_rango.empty:
            df_rangos_list.append(df_rango)
    df_rangos = pd.concat(df_rangos_list) if df_rangos_list else pd.DataFrame()
    huella_min_rangos = df_rangos['huella_co2'].min() if not df_rangos.empty else 0
    huella_max_rangos = df_rangos['huella_co2'].max() if not df_rangos.empty else 0
    huella_prom_rangos = df_rangos['huella_co2'].mean() if not df_rangos.empty else 0
    volumen_rangos = df_rangos['volumen'].sum() if 'volumen' in df_rangos.columns and not df_rangos.empty else 0

    # Escribir métricas
    metricas = [
        ['Huella CO₂ Mínima (kg/m³)', huella_min_todas, huella_min_exactos, huella_min_rangos],
        ['Huella CO₂ Máxima (kg/m³)', huella_max_todas, huella_max_exactos, huella_max_rangos],
        ['Huella CO₂ Promedio (kg/m³)', huella_prom_todas, huella_prom_exactos, huella_prom_rangos],
        ['Volumen Total (m³)', volumen_todas, volumen_exactos, volumen_rangos]
    ]

    for row_idx, metrica_data in enumerate(metricas, start=10):
        for col_idx, valor in enumerate(metrica_data, start=1):
            cell = ws_indice.cell(row=row_idx, column=col_idx, value=valor)
            cell.border = border
            if col_idx > 1:
                cell.number_format = '0.00' if row_idx <= 12 else '0'
            cell.alignment = Alignment(horizontal='center')

    # Ajustar anchos
    ws_indice.column_dimensions['A'].width = 30
    ws_indice.column_dimensions['B'].width = 15
    ws_indice.column_dimensions['C'].width = 15
    ws_indice.column_dimensions['D'].width = 15

    # HOJA 2: Todas las resistencias
    ws_todas = wb.create_sheet("Todas las Resistencias")
    ws_todas['A1'] = "Resistencia (MPa)"
    ws_todas['B1'] = "Huella CO₂ (kg/m³)"
    ws_todas['C1'] = "Volumen (m³)"

    for col in ['A1', 'B1', 'C1']:
        ws_todas[col].fill = header_fill
        ws_todas[col].font = header_font
        ws_todas[col].border = border
        ws_todas[col].alignment = Alignment(horizontal='center')

    df_todas_sorted = df_todas.sort_values('REST')
    for row_idx, (_, row) in enumerate(df_todas_sorted.iterrows(), start=2):
        ws_todas.cell(row=row_idx, column=1, value=row['REST']).border = border
        ws_todas.cell(row=row_idx, column=2, value=row['huella_co2']).border = border
        ws_todas.cell(row=row_idx, column=2).number_format = '0.00'
        if 'volumen' in row:
            ws_todas.cell(row=row_idx, column=3, value=row['volumen']).border = border
            ws_todas.cell(row=row_idx, column=3).number_format = '0'

    ws_todas.column_dimensions['A'].width = 20
    ws_todas.column_dimensions['B'].width = 20
    ws_todas.column_dimensions['C'].width = 15

    # HOJA 3: Solo exactos
    ws_exactos = wb.create_sheet("Solo Exactos")
    ws_exactos['A1'] = "Resistencia (MPa)"
    ws_exactos['B1'] = "Huella CO₂ (kg/m³)"
    ws_exactos['C1'] = "Volumen (m³)"

    for col in ['A1', 'B1', 'C1']:
        ws_exactos[col].fill = header_fill
        ws_exactos[col].font = header_font
        ws_exactos[col].border = border
        ws_exactos[col].alignment = Alignment(horizontal='center')

    df_exactos_sorted = df_exactos.sort_values('REST') if not df_exactos.empty else pd.DataFrame()
    for row_idx, (_, row) in enumerate(df_exactos_sorted.iterrows(), start=2):
        ws_exactos.cell(row=row_idx, column=1, value=row['REST']).border = border
        ws_exactos.cell(row=row_idx, column=2, value=row['huella_co2']).border = border
        ws_exactos.cell(row=row_idx, column=2).number_format = '0.00'
        if 'volumen' in row:
            ws_exactos.cell(row=row_idx, column=3, value=row['volumen']).border = border
            ws_exactos.cell(row=row_idx, column=3).number_format = '0'

    ws_exactos.column_dimensions['A'].width = 20
    ws_exactos.column_dimensions['B'].width = 20
    ws_exactos.column_dimensions['C'].width = 15

    # HOJA 4: Rangos ±2 MPa
    ws_rangos = wb.create_sheet("Rangos ±2 MPa")
    ws_rangos['A1'] = "Resistencia (MPa)"
    ws_rangos['B1'] = "Huella CO₂ (kg/m³)"
    ws_rangos['C1'] = "Volumen (m³)"

    for col in ['A1', 'B1', 'C1']:
        ws_rangos[col].fill = header_fill
        ws_rangos[col].font = header_font
        ws_rangos[col].border = border
        ws_rangos[col].alignment = Alignment(horizontal='center')

    df_rangos_sorted = df_rangos.sort_values('REST') if not df_rangos.empty else pd.DataFrame()
    for row_idx, (_, row) in enumerate(df_rangos_sorted.iterrows(), start=2):
        ws_rangos.cell(row=row_idx, column=1, value=row['REST']).border = border
        ws_rangos.cell(row=row_idx, column=2, value=row['huella_co2']).border = border
        ws_rangos.cell(row=row_idx, column=2).number_format = '0.00'
        if 'volumen' in row:
            ws_rangos.cell(row=row_idx, column=3, value=row['volumen']).border = border
            ws_rangos.cell(row=row_idx, column=3).number_format = '0'

    ws_rangos.column_dimensions['A'].width = 20
    ws_rangos.column_dimensions['B'].width = 20
    ws_rangos.column_dimensions['C'].width = 15

    # Guardar en BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output

def app():
    # Obtener conexión de session_state
    ruta_db = st.session_state.get('ruta_db')
    conn = get_connection(ruta_db)

    st.title("📋 Bandas GCCA Concretos")

    # ==================== MODO FANTASMA ====================
    st.sidebar.header("👻 Modo Fantasma")
    modo_fantasma = st.sidebar.checkbox("Modo Fantasma", value=True)

    # Cargar bandas
    json_path = os.path.join(os.getenv("COMUN_FILES_PATH"), "bandas_gcca.json")
    bandas = cargar_bandas(json_path)

    # Cargar datos desde la base de datos
    df_csv = pd.read_sql_query("SELECT * FROM huella_concretos", conn)

    # Crear mapeo de nombres reales a anónimos (alfabéticamente ordenados)
    origenes_unicos = sorted(df_csv['origen'].unique())
    mapeo_fantasma = {origen: f"Compañía {i+1}" for i, origen in enumerate(origenes_unicos)}

    # Aplicar modo fantasma si está activado
    if modo_fantasma:
        df_csv['origen'] = df_csv['origen'].map(mapeo_fantasma)

    df_csv["serie"] = df_csv["año"].astype(str) + " - " + df_csv["origen"]

    # Selector múltiple
    series_seleccionadas = st.multiselect(
        "Selecciona series a mostrar (Año - Origen):",
        options=df_csv["serie"].unique(),
        default=[]
    )

    # NO seguir si no hay series
    if len(series_seleccionadas) == 0:
        st.stop()

    # Filtrar datos por series seleccionadas
    df_filtrado = df_csv[df_csv["serie"].isin(series_seleccionadas)]

    # ESTRUCTURA DE PESTAÑAS
    tab1, tab2, tab3, tab4 = st.tabs(["Visualización Básica", "Análisis Integrado", "Bandas GCCA + Datos", "Análisis por Compañía"])

    # ==================== TAB 1: VISUALIZACIÓN BÁSICA ====================
    with tab1:
        # Selector de tipo de visualización
        tipo_visualizacion = st.radio(
            "Selecciona el tipo de visualización:",
            ["Gráfico de Líneas", "Gráfico de Burbujas", "Gráfico de Barras Apiladas"],
            horizontal=True
        )

        # Asegurarse de que bandas esté ordenado
        df_bandas = pd.DataFrame.from_dict(bandas, orient='index')
        df_bandas = df_bandas[[20, 25, 30, 35, 40, 50]]
        df_bandas = df_bandas.sort_values(by=20)

        # Definir un valor máximo para la banda H
        valor_maximo_banda_h = 800

        # Añadir la banda H
        if "Top of F" in df_bandas.index:
            banda_h = pd.Series(
                data=[valor_maximo_banda_h] * len(df_bandas.columns),
                index=df_bandas.columns
            )
            df_bandas.loc["Top of H"] = banda_h

        # Calcular valores por fila
        bandas_ordenadas = df_bandas.index.tolist()
        colores_bandas = [
            "#2ca02c",  # AA Near Zero
            "#006400",  # A
            "#1f77b4",  # B
            "#00bfff",  # C
            "#007acc",  # D
            "#7f7f7f",  # E
            "#4d4d4d",  # F
            "#1f1f1f",  # H
        ]

        columnas_remitos = {
            "Top of AA -Near Zero Product": "remAA_Near_Zero_Product",
            "Top of A": "A",
            "Top of B": "B",
            "Top of C": "C",
            "Top of D": "D",
            "Top of E": "E",
            "Top of F": "F",
            "Top of H": "H"
        }

        columnas_disponibles = [col for col in df_filtrado.columns if col.startswith("")]

        # Función auxiliar para clasificar valores en bandas
        def clasificar_en_bandas(rest, huella, df_bandas):
            for banda in bandas_ordenadas:
                if rest in df_bandas.columns and huella <= df_bandas.loc[banda, rest]:
                    return banda
            return "Top of H"

        # Preprocesar datos para clasificar remitos en la banda H
        for serie in series_seleccionadas:
            df_serie = df_filtrado[df_filtrado["serie"] == serie]

            if "H" not in df_filtrado.columns:
                df_filtrado["H"] = 0

            for idx, row in df_serie.iterrows():
                rest = row["REST"]
                huella = row["huella_co2"]
                banda = clasificar_en_bandas(rest, huella, df_bandas)

                if banda == "Top of H":
                    remitos_en_otras_bandas = sum(row.get(columnas_remitos.get(b, ""), 0) for b in bandas_ordenadas if b != "Top of H")
                    if remitos_en_otras_bandas == 0 and row.get("remitos", 0) > 0:
                        df_filtrado.at[idx, "H"] = row.get("remitos", 0)

        if tipo_visualizacion == "Gráfico de Líneas":
            # Gráfico original de líneas con bandas
            fig, ax = plt.subplots(figsize=(12, 7))
            x = df_bandas.columns.astype(int)

            # Dibujar bandas
            base_inferior = pd.Series([0]*len(x), index=x)
            for i, banda in enumerate(bandas_ordenadas):
                top = df_bandas.loc[banda]
                ax.fill_between(
                    x, base_inferior, top,
                    label=banda.replace("Top of ", "").replace(" -Near Zero Product", " Near Zero"),
                    color=colores_bandas[i],
                    alpha=0.7
                )
                base_inferior = top

            # Agregar series seleccionadas
            for serie in series_seleccionadas:
                df_serie = df_filtrado[df_filtrado["serie"] == serie]
                df_serie_ordenada = df_serie.sort_values(by="REST")

                ax.plot(
                    df_serie_ordenada["REST"],
                    df_serie_ordenada["huella_co2"],
                    marker="o",
                    label=serie,
                    linewidth=2
                )

                # Marcar específicamente los puntos que caen en la banda H
                for _, row in df_serie_ordenada.iterrows():
                    rest = row["REST"]
                    huella = row["huella_co2"]
                    banda = clasificar_en_bandas(rest, huella, df_bandas)

                    if banda == "Top of H" and row.get("H", 0) > 0:
                        ax.scatter(
                            rest, huella,
                            s=100,
                            color=colores_bandas[bandas_ordenadas.index("Top of H")],
                            marker="*",
                            edgecolor="black",
                            linewidth=1,
                            zorder=5
                        )

            ax.set_title("Bandas GCCA + Series seleccionadas")
            ax.set_xlabel("Resistencia (MPa)")
            ax.set_ylabel("Huella de CO₂ (kg/m³)")
            ax.set_ylim(0, valor_maximo_banda_h)
            ax.set_xlim(min(x), max(x))
            ax.legend(loc="upper left", fontsize=8)
            ax.grid(True, linestyle="--", alpha=0.5)

        elif tipo_visualizacion == "Gráfico de Burbujas":
            # Visualización de burbujas
            fig, ax = plt.subplots(figsize=(14, 8))
            x = df_bandas.columns.astype(int)

            # Dibujar bandas con transparencia
            base_inferior = pd.Series([0]*len(x), index=x)
            for i, banda in enumerate(bandas_ordenadas):
                top = df_bandas.loc[banda]
                ax.fill_between(
                    x, base_inferior, top,
                    label=banda.replace("Top of ", "").replace(" -Near Zero Product", " Near Zero"),
                    color=colores_bandas[i],
                    alpha=0.3
                )
                base_inferior = top

            banda_cmap = LinearSegmentedColormap.from_list('banda_cmap', colores_bandas)

            for serie_idx, serie in enumerate(series_seleccionadas):
                df_serie = df_filtrado[df_filtrado["serie"] == serie]

                for idx, row in df_serie.iterrows():
                    rest = row["REST"]
                    huella = row["huella_co2"]
                    banda_actual = clasificar_en_bandas(rest, huella, df_bandas)

                    for i, banda in enumerate(bandas_ordenadas):
                        nombre_banda = banda.replace("Top of ", "").replace(" -Near Zero Product", " Near Zero")
                        col_nombre = columnas_remitos.get(banda, "")

                        if col_nombre in columnas_disponibles and col_nombre in row:
                            remitos = row[col_nombre]
                            if remitos > 0:
                                size = np.sqrt(remitos) * 20

                                if banda == "Top of H":
                                    marker = '*'
                                    edge_width = 1.5
                                else:
                                    marker = 'o'
                                    edge_width = 0.5

                                ax.scatter(
                                    rest, huella,
                                    s=size,
                                    color=colores_bandas[i],
                                    alpha=0.7,
                                    edgecolor='black',
                                    linewidth=edge_width,
                                    marker=marker,
                                    label=f"{serie} - {nombre_banda}" if serie_idx == 0 and i == 0 else ""
                                )

                                ax.annotate(
                                    f"{int(remitos)}",
                                    (rest, huella),
                                    textcoords="offset points",
                                    xytext=(0, 0),
                                    ha='center',
                                    va='center',
                                    fontsize=8,
                                    color='white',
                                    weight='bold'
                                )

            ax.set_title("Dispersión de Remitos por Banda y Resistencia")
            ax.set_xlabel("Resistencia (MPa)")
            ax.set_ylabel("Huella de CO₂ (kg/m³)")
            ax.set_xlim(min(x) - 2, max(x) + 2)
            ax.set_ylim(0, valor_maximo_banda_h)
            ax.grid(True, linestyle="--", alpha=0.5)

            import matplotlib.patches as mpatches
            patches = []
            for i, banda in enumerate(bandas_ordenadas):
                nombre = banda.replace("Top of ", "").replace(" -Near Zero Product", " Near Zero")
                patch = mpatches.Patch(color=colores_bandas[i], label=nombre)
                patches.append(patch)

            ax.legend(handles=patches, loc="upper left", fontsize=8)

        else:  # Gráfico de Barras Apiladas
            if not series_seleccionadas:
                st.warning("Selecciona al menos una serie para mostrar el gráfico de barras apiladas.")
            else:
                n_series = len(series_seleccionadas)
                fig, axes = plt.subplots(n_series, 1, figsize=(12, 5 * n_series), sharex=True)

                if n_series == 1:
                    axes = [axes]

                resistencias = sorted(df_bandas.columns.astype(int))

                for i, serie in enumerate(series_seleccionadas):
                    ax = axes[i]
                    df_serie = df_filtrado[df_filtrado["serie"] == serie]

                    datos_barras = []
                    for rest in resistencias:
                        row = df_serie[df_serie["REST"] == rest]
                        if not row.empty:
                            datos = [0] * len(bandas_ordenadas)
                            for j, banda in enumerate(bandas_ordenadas):
                                col_nombre = columnas_remitos.get(banda, "")
                                if col_nombre in columnas_disponibles and col_nombre in row.columns:
                                    datos[j] = row[col_nombre].values[0]
                            datos_barras.append(datos)
                        else:
                            datos_barras.append([0] * len(bandas_ordenadas))

                    datos_barras = np.array(datos_barras).T

                    bottom = np.zeros(len(resistencias))
                    for j, valores in enumerate(datos_barras):
                        nombre_banda = bandas_ordenadas[j].replace("Top of ", "").replace(" -Near Zero Product", " Near Zero")
                        ax.bar(
                            resistencias,
                            valores,
                            bottom=bottom,
                            label=nombre_banda,
                            color=colores_bandas[j],
                            edgecolor='white',
                            linewidth=0.5
                        )

                        for k, v in enumerate(valores):
                            if v > 0:
                                ax.text(
                                    resistencias[k],
                                    bottom[k] + v/2,
                                    str(int(v)),
                                    ha='center',
                                    va='center',
                                    color='white',
                                    fontweight='bold',
                                    fontsize=9
                                )
                        bottom += valores

                    ax.set_title(f"Distribución de Remitos por Banda - {serie}")
                    ax.set_ylabel("Número de Remitos")
                    ax.grid(True, axis='y', linestyle='--', alpha=0.7)

                    if i == 0:
                        ax.legend(loc='upper right', fontsize=8)

                axes[-1].set_xlabel("Resistencia (MPa)")
                plt.tight_layout()

        st.pyplot(fig)

        # Tabla resumen opcional
        if st.checkbox("Mostrar tabla resumen de remitos por banda"):
            for serie in series_seleccionadas:
                st.subheader(f"Resumen para {serie}")

                df_resumen = df_filtrado[df_filtrado["serie"] == serie]

                tabla_resumen = []
                for _, row in df_resumen.iterrows():
                    datos_fila = {"Resistencia (MPa)": row["REST"], "Huella CO₂ (kg/m³)": row["huella_co2"]}

                    for banda in bandas_ordenadas:
                        col_nombre = columnas_remitos.get(banda, "")
                        if col_nombre in columnas_disponibles and col_nombre in row:
                            nombre_banda = banda.replace("Top of ", "").replace(" -Near Zero Product", " Near Zero")
                            datos_fila[f"Remitos {nombre_banda}"] = row[col_nombre]

                    tabla_resumen.append(datos_fila)

                if tabla_resumen:
                    df_tabla = pd.DataFrame(tabla_resumen)
                    st.dataframe(df_tabla.set_index("Resistencia (MPa)"))

    # ==================== TAB 2: ANÁLISIS INTEGRADO ====================
    with tab2:
        st.subheader("Análisis Integrado por Resistencia")

        resistencias_estandar = [20, 25, 30, 35, 40, 45, 50]

        # Selector de modo de filtrado
        modo_resistencia = st.radio(
            "⚙️ Selecciona el modo de análisis:",
            [
                "Todas las resistencias disponibles",
                f"Solo resistencias exactas ({', '.join(map(str, resistencias_estandar))} MPa)",
                "Rangos ±2 MPa alrededor de resistencias estándar"
            ],
            index=0
        )

        st.divider()

        # Aplicar filtrado según modo seleccionado
        if modo_resistencia.startswith("Todas"):
            df_analisis = df_filtrado.copy()
            modo_key = "todas"
            modo_descripcion = "Incluye todas las resistencias sin filtrado"
        elif modo_resistencia.startswith("Solo"):
            df_analisis = df_filtrado[df_filtrado['REST'].isin(resistencias_estandar)]
            modo_key = "exactos"
            modo_descripcion = f"Solo resistencias exactas: {', '.join(map(str, resistencias_estandar))} MPa"
        else:  # Rangos
            df_rangos_list = []
            for res_std in resistencias_estandar:
                df_rango = df_filtrado[(df_filtrado['REST'] >= res_std - 2) & (df_filtrado['REST'] <= res_std + 2)]
                if not df_rango.empty:
                    df_rangos_list.append(df_rango)
            df_analisis = pd.concat(df_rangos_list) if df_rangos_list else pd.DataFrame()
            modo_key = "rangos"
            modo_descripcion = "Agrupa resistencias en rangos de ±2 MPa alrededor de valores estándar"

        if df_analisis.empty:
            st.warning("No hay datos disponibles con el filtrado seleccionado.")
            st.stop()

        # MANOMETERS (4 métricas)
        col1, col2, col3, col4 = st.columns(4)

        with col1:
            huella_min = df_analisis['huella_co2'].min()
            fig_gauge1 = go.Figure(go.Indicator(
                mode="gauge+number",
                value=huella_min,
                title={'text': "Huella CO₂ Mínima"},
                gauge={'axis': {'range': [0, 800]},
                       'bar': {'color': "#2ca02c"},
                       'threshold': {
                           'line': {'color': "red", 'width': 4},
                           'thickness': 0.75,
                           'value': 400
                       }}
            ))
            fig_gauge1.update_layout(height=250)
            st.plotly_chart(fig_gauge1, use_container_width=True)

        with col2:
            huella_max = df_analisis['huella_co2'].max()
            fig_gauge2 = go.Figure(go.Indicator(
                mode="gauge+number",
                value=huella_max,
                title={'text': "Huella CO₂ Máxima"},
                gauge={'axis': {'range': [0, 800]},
                       'bar': {'color': "#d62728"},
                       'threshold': {
                           'line': {'color': "red", 'width': 4},
                           'thickness': 0.75,
                           'value': 400
                       }}
            ))
            fig_gauge2.update_layout(height=250)
            st.plotly_chart(fig_gauge2, use_container_width=True)

        with col3:
            huella_prom = df_analisis['huella_co2'].mean()
            fig_gauge3 = go.Figure(go.Indicator(
                mode="gauge+number",
                value=huella_prom,
                title={'text': "Huella CO₂ Promedio"},
                gauge={'axis': {'range': [0, 800]},
                       'bar': {'color': "#1f77b4"},
                       'threshold': {
                           'line': {'color': "red", 'width': 4},
                           'thickness': 0.75,
                           'value': 400
                       }}
            ))
            fig_gauge3.update_layout(height=250)
            st.plotly_chart(fig_gauge3, use_container_width=True)

        with col4:
            volumen_total = df_analisis['volumen'].sum() if 'volumen' in df_analisis.columns else 0
            fig_gauge4 = go.Figure(go.Indicator(
                mode="number",
                value=volumen_total,
                title={'text': "Volumen Total (m³)"},
                number={'suffix': " m³", 'valueformat': ",.0f"}
            ))
            fig_gauge4.update_layout(height=250)
            st.plotly_chart(fig_gauge4, use_container_width=True)

        st.divider()

        # GRÁFICO DE VOLUMEN + INTENSIDAD CON BANDAS
        st.subheader("Volumen de Producción e Intensidad de Emisión por Resistencia")

        # Agrupar por resistencia
        if 'volumen' in df_analisis.columns:
            df_agrupado = df_analisis.groupby('REST').agg({
                'volumen': 'sum',
                'huella_co2': 'mean'
            }).reset_index()
            df_agrupado = df_agrupado.sort_values('REST')

            # Crear figura con dos ejes Y
            fig2, ax1 = plt.subplots(figsize=(14, 7))

            # Eje 1: Barras de volumen
            color_barras = '#1f77b4'
            ax1.set_xlabel('Resistencia (MPa)', fontsize=12)
            ax1.set_ylabel('Volumen de Producción (m³)', color=color_barras, fontsize=12)
            ax1.bar(df_agrupado['REST'], df_agrupado['volumen'],
                   color=color_barras, alpha=0.6, label='Volumen Producción')
            ax1.tick_params(axis='y', labelcolor=color_barras)
            ax1.grid(True, axis='y', linestyle='--', alpha=0.3)

            # Agregar etiquetas en las barras de volumen
            for rest, vol in zip(df_agrupado['REST'], df_agrupado['volumen']):
                ax1.text(rest, vol, f'{int(vol):,}',
                        ha='center', va='bottom', fontsize=9, fontweight='bold',
                        color=color_barras)

            # Eje 2: Línea de huella CO2 con bandas
            ax2 = ax1.twinx()
            ax2.set_ylabel('Huella de CO₂ (kg/m³)', color='#d62728', fontsize=12)

            # Dibujar bandas GCCA en el eje secundario
            x_bandas = df_bandas.columns.astype(int)
            base_inferior = pd.Series([0]*len(x_bandas), index=x_bandas)
            for i, banda in enumerate(bandas_ordenadas):
                top = df_bandas.loc[banda]
                ax2.fill_between(
                    x_bandas, base_inferior, top,
                    color=colores_bandas[i],
                    alpha=0.35,
                    zorder=1
                )
                base_inferior = top

            # Línea de huella CO2 adelante
            ax2.plot(df_agrupado['REST'], df_agrupado['huella_co2'],
                    color='#d62728', marker='o', linewidth=2.5,
                    markersize=8, label='Huella CO₂ Promedio', zorder=5)
            ax2.tick_params(axis='y', labelcolor='#d62728')
            ax2.set_ylim(0, 800)

            # Agregar etiquetas en los puntos de huella CO2
            for rest, huella in zip(df_agrupado['REST'], df_agrupado['huella_co2']):
                ax2.text(rest, huella + 15, f'{huella:.1f}',
                        ha='center', va='bottom', fontsize=9, fontweight='bold',
                        color='#d62728', bbox=dict(boxstyle='round,pad=0.3',
                                                   facecolor='white', alpha=0.8))

            fig2.tight_layout()
            plt.title('Volumen de Producción e Intensidad de Emisión por Resistencia',
                     fontsize=14, fontweight='bold', pad=20)

            # Cuadro descriptivo
            fig2.text(0.5, -0.08,
                     f"⚙️ Opción seleccionada: {modo_resistencia.split('(')[0].strip()}\n📝 {modo_descripcion}",
                     ha='center', va='top', fontsize=10,
                     bbox=dict(boxstyle='round,pad=0.8', facecolor='#F0F0F0',
                              edgecolor='#333333', linewidth=1.5))

            st.pyplot(fig2)

            # Botón de descarga del gráfico
            nombres_archivo = {
                "todas": "grafico_volumen_intensidad_todas.png",
                "exactos": "grafico_volumen_intensidad_exactos.png",
                "rangos": "grafico_volumen_intensidad_rangos.png"
            }

            buf = BytesIO()
            fig2.savefig(buf, format='png', dpi=300, bbox_inches='tight')
            buf.seek(0)

            st.download_button(
                label="📥 Descargar Gráfico",
                data=buf,
                file_name=nombres_archivo[modo_key],
                mime="image/png"
            )

            st.divider()

            # Tabla de datos integrados
            st.subheader("Tabla de Datos Integrados")
            st.dataframe(df_agrupado)

        # Botón de descarga de Excel
        excel_data = generar_excel_analisis(df_filtrado, bandas)
        st.download_button(
            label="📥 Descargar Análisis Completo (Excel)",
            data=excel_data,
            file_name="analisis_bandas_concretos_completo.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    # ==================== TAB 3: BANDAS GCCA + DATOS ====================
    with tab3:
        st.subheader("Bandas GCCA + Datos Integrados de Todas las Series")

        # Gráfico con bandas y todas las series
        fig3, ax3 = plt.subplots(figsize=(14, 8))
        x = df_bandas.columns.astype(int)

        # Dibujar bandas
        base_inferior = pd.Series([0]*len(x), index=x)
        for i, banda in enumerate(bandas_ordenadas):
            top = df_bandas.loc[banda]
            ax3.fill_between(
                x, base_inferior, top,
                label=banda.replace("Top of ", "").replace(" -Near Zero Product", " Near Zero"),
                color=colores_bandas[i],
                alpha=0.7
            )
            base_inferior = top

        # Agregar todas las series seleccionadas
        for serie in series_seleccionadas:
            df_serie = df_filtrado[df_filtrado["serie"] == serie]
            df_serie_ordenada = df_serie.sort_values(by="REST")

            ax3.plot(
                df_serie_ordenada["REST"],
                df_serie_ordenada["huella_co2"],
                marker="o",
                label=serie,
                linewidth=2
            )

        ax3.set_title("Bandas GCCA + Todas las Series Seleccionadas", fontsize=14, fontweight='bold')
        ax3.set_xlabel("Resistencia (MPa)", fontsize=12)
        ax3.set_ylabel("Huella de CO₂ (kg/m³)", fontsize=12)
        ax3.set_ylim(0, valor_maximo_banda_h)
        ax3.set_xlim(min(x), max(x))
        ax3.legend(loc="upper left", fontsize=9)
        ax3.grid(True, linestyle="--", alpha=0.5)

        st.pyplot(fig3)

    # ==================== TAB 4: ANÁLISIS POR COMPAÑÍA ====================
    with tab4:
        st.subheader("📊 Análisis Comparativo por Compañía")

        # Verificar que hay datos de múltiples compañías
        if df_filtrado['origen'].nunique() > 1:

            # ========== SECCIÓN 1: Gráfico de Torta ==========
            st.markdown("### Distribución de Volumen por Compañía")

            # Agrupar por compañía
            df_por_compania = df_filtrado.groupby('origen').agg({
                'volumen': 'sum',
                'huella_co2': 'mean'
            }).reset_index()

            # Gráfico de torta
            fig_pie = px.pie(
                df_por_compania,
                values='volumen',
                names='origen',
                title='Distribución del Volumen Total por Compañía',
                hole=0.3,
                color_discrete_sequence=px.colors.qualitative.Set3
            )

            fig_pie.update_traces(
                textposition='inside',
                textinfo='percent+label',
                hovertemplate='<b>%{label}</b><br>Volumen: %{value:,.0f} m³<br>Porcentaje: %{percent}<extra></extra>'
            )

            st.plotly_chart(fig_pie, use_container_width=True)

            st.divider()

            # ========== SECCIÓN 2: Gráfico de Barras Agrupadas ==========
            st.markdown("### Volumen de Producción por Compañía y Año")

            # Agrupar por compañía y año
            df_compania_ano = df_filtrado.groupby(['origen', 'año']).agg({
                'volumen': 'sum',
                'huella_co2': 'mean'
            }).reset_index()

            # Gráfico de barras agrupadas
            fig_bar = px.bar(
                df_compania_ano,
                x='año',
                y='volumen',
                color='origen',
                barmode='group',
                title='Volumen de Producción por Compañía y Año',
                labels={
                    'volumen': 'Volumen (m³)',
                    'año': 'Año',
                    'origen': 'Compañía'
                },
                color_discrete_sequence=px.colors.qualitative.Bold,
                text='volumen'
            )

            fig_bar.update_traces(
                texttemplate='%{text:,.0f}',
                textposition='outside',
                hovertemplate='<b>%{fullData.name}</b><br>Año: %{x}<br>Volumen: %{y:,.0f} m³<extra></extra>'
            )

            fig_bar.update_layout(
                xaxis=dict(
                    tickmode='linear',
                    tick0=df_compania_ano['año'].min(),
                    dtick=1
                ),
                yaxis_title='Volumen (m³)',
                legend_title='Compañía',
                hovermode='x unified',
                height=500
            )

            st.plotly_chart(fig_bar, use_container_width=True)

            st.divider()

            # ========== SECCIÓN 3: Tabla Resumen ==========
            st.markdown("### Tabla Resumen por Compañía y Año")

            # Preparar tabla con formato
            df_tabla = df_compania_ano.copy()
            df_tabla = df_tabla.rename(columns={
                'origen': 'Compañía',
                'año': 'Año',
                'volumen': 'Volumen (m³)',
                'huella_co2': 'Huella CO₂ Promedio (kg/m³)'
            })

            st.dataframe(
                df_tabla,
                use_container_width=True,
                hide_index=True,
                column_config={
                    'Volumen (m³)': st.column_config.NumberColumn(format="%.0f"),
                    'Huella CO₂ Promedio (kg/m³)': st.column_config.NumberColumn(format="%.2f")
                }
            )

            # ========== SECCIÓN 4: Estadísticas por Compañía ==========
            st.divider()
            st.markdown("### Estadísticas por Compañía")

            cols = st.columns(len(df_por_compania))

            for idx, (_, row) in enumerate(df_por_compania.iterrows()):
                with cols[idx]:
                    st.metric(
                        label=row['origen'],
                        value=f"{row['volumen']:,.0f} m³",
                        delta=f"Huella: {row['huella_co2']:.2f} kg/m³"
                    )

            # ========== SECCIÓN 5: Botón de Descarga ==========
            st.divider()

            # Generar Excel para descarga
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_compania_ano.to_excel(writer, sheet_name='Por Compañía y Año', index=False)
                df_por_compania.to_excel(writer, sheet_name='Resumen por Compañía', index=False)
            output.seek(0)

            st.download_button(
                label="📥 Descargar Análisis por Compañía (Excel)",
                data=output,
                file_name="analisis_compania_bandas_concretos.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        else:
            st.warning("⚠️ No hay suficientes compañías en los datos seleccionados para realizar el análisis comparativo.")
